import openpyxl
import smtplib


# Сохраняем изменения
def save():
    book.save(r'C:\\Users\\andrs\\PycharmProjects\\project of attendance\\attendance.xlsx')
    print("saved!")


# Записываем отсутствующих
def lack_of_attendance(student, subject):
    num = int(book['Лист1'].cell(student, subject).value)
    num += 1
    book['Лист1'].cell(student, subject).value = num
    save()


# Обнуление посещаемости
def reset():
    for a in range(2, sheet.max_row + 1):
        for b in range(3, sheet.max_column + 1):
            sheet.cell(a, b).value = 0
    save()


# Узнаем "лидера" по посещаемости
def leader_search():
    global leader
    maxk = 0

    for a in range(2, sheet.max_row + 1):
        for b in range(3, sheet.max_column + 1):
            if int(sheet.cell(a, b).value) > maxk:
                maxk = int(sheet.cell(a, b).value)
                leader = sheet.cell(a, 1).value

    return leader


# Метод для рассылки писем
def send_notification(email, txt):
    sender = '****'
    sender_password = '****'
    mail_lib = smtplib.SMTP_SSL('smtp.yandex.ru', 465)
    mail_lib.login(sender, sender_password)
    for to_item in email:
        msg = 'From: %s\r\nTo: %s\r\nContent-Type: text/plain; charset="utf-8"\r\nSubject: %s\r\n\r\n' % (
            sender, to_item, 'Предупреждение')
        msg += txt
        mail_lib.sendmail(sender, to_item, msg.encode('utf8'))
    mail_lib.quit()


# Метод для поиска кандидатов с превышением допустимых пропусков
def candidate_search():
    candidates = []
    for a in range(2, sheet.max_row + 1):
        for b in range(3, sheet.max_column + 1):
            if int(sheet.cell(a, b).value) > 2:
                candidates.append(sheet.cell(a, 2).value)
    return candidates


# Создание переменной с файлом Excel
book = openpyxl.load_workbook('C:\\Users\\andrs\\PycharmProjects\\project of attendance\\attendance.xlsx')

# Создание переменной страницы
sheet = book['Лист1']

# Примерные сообщения для отправки
msg = 'Внимание, Вы кандидат на вылет!'

print('Есть ли сегодня отстутствующие?')
check_of_attendance = 1

# Узнаем отсутствующих
while check_of_attendance == 1:

    check_of_attendance = int(input('1 -- Да\n2 -- Нет\n'))
    if check_of_attendance == 1:

        # Узнаем название предмета
        for i in range(3, sheet.max_column + 1):
            print(str(i - 2) + '. ' + sheet.cell(1, i).value)
        subject_id = input('Введите ID предмета:\n')

        # Узнаем ID ученика
        for i in range(2, sheet.max_row + 1):
            print(str(i - 1) + '. ' + sheet.cell(i, 1).value)
        student_id = input('Введите ID ученика:\n')
        lack_of_attendance(student_id, subject_id)
        print('Есть ли еще?\n')

print('--- Основное меню ---\n')

menu: str = input('1. Узнать дежурного\n2. Обнулить посещаемость\n3. Разослать e-mail с предупреждением\n')

if menu == '1':
    print('В этом месяце дежурит: ' + leader_search())

elif menu == '2':
    reset()

else:
    email = candidate_search()
    send_notification(email, msg)
